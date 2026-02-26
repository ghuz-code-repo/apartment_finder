import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.core.db_utils import get_planning_session, get_mysql_session
from app.models import planning_models
from app.models.estate_models import EstateSell, EstateHouse
from app.models.planning_models import map_russian_to_mysql_key, PropertyType
from app.services import currency_service

RESERVATION_FEE = 3_000_000
REMAINDER_STATUSES = ["Маркетинговый резерв", "Подбор"]


def calculate_new_prices(complex_name, property_type_ru, percent_change, excluded_ids=None):
    """
    Рассчитывает новые цены с логикой перераспределения повышения.
    Сумма повышения, приходящаяся на исключенные объекты, равномерно распределяется на остальные.
    Возвращает: (technical_results, stats_with_action, stats_no_action)
    """
    if excluded_ids is None:
        excluded_ids = []

    mysql_session = get_mysql_session()
    planning_session = get_planning_session()
    try:
        usd_rate = currency_service.get_current_effective_rate() or 12800.0
        prop_type_enum = PropertyType(property_type_ru)
        mysql_key = map_russian_to_mysql_key(property_type_ru)

        active_version = planning_session.query(planning_models.DiscountVersion).filter_by(is_active=True).first()
        if not active_version:
            return None, None, "Нет активной версии скидок"

        discount = planning_session.query(planning_models.Discount).filter_by(
            version_id=active_version.id,
            complex_name=complex_name,
            property_type=prop_type_enum,
            payment_method=planning_models.PaymentMethod.FULL_PAYMENT
        ).first()

        # Расчет коэффициентов скидок
        total_rate_no_action = 0
        total_rate_with_action = 0
        if discount:
            total_rate_no_action = (discount.mpp or 0) + (discount.rop or 0) + (discount.kd or 0)
            total_rate_with_action = total_rate_no_action + (discount.action or 0)

        mult_with = 1 - total_rate_with_action
        mult_no = 1 - total_rate_no_action
        current_res_fee = RESERVATION_FEE if property_type_ru == 'Квартира' else 0

        all_objects = mysql_session.query(EstateSell).filter(
            EstateSell.house_id.in_(
                mysql_session.query(EstateHouse.id).filter_by(complex_name=complex_name)
            ),
            EstateSell.estate_sell_category == mysql_key
        ).all()

        # --- ЛОГИКА РАЗМЫТИЯ (РАСЧЕТ ЭФФЕКТИВНОГО ПРОЦЕНТА) ---
        sum_net_all = 0.0
        sum_net_non_excluded = 0.0

        for obj in all_objects:
            if not obj.estate_price or not obj.estate_area or obj.estate_area <= 0:
                continue
            net_val = (obj.estate_price - current_res_fee)
            sum_net_all += net_val
            if obj.id not in excluded_ids:
                sum_net_non_excluded += net_val

        if sum_net_non_excluded > 0:
            adjusted_percent = percent_change * (sum_net_all / sum_net_non_excluded)
        else:
            adjusted_percent = percent_change

        # --- ИНИЦИАЛИЗАЦИЯ СТАТИСТИКИ ДЛЯ ДВУХ ВАРИАНТОВ ---
        def init_stats(discount_rate, base_pct, adj_pct):
            return {
                'complex_name': complex_name, 'prop_type': property_type_ru, 'usd_rate': usd_rate,
                'total_units_project': len(all_objects), 'remainder_units': 0, 'remainder_sqm': 0.0,
                'discount_pct': round(discount_rate * 100, 2),
                'percent_change': round(base_pct * 100, 2),
                'adjusted_percent': round(adj_pct * 100, 2),
                'floor_prices_before': [], 'floor_prices_after': [],
                'floor_totals_before': [], 'floor_totals_after': [],
                'final_totals_before': 0.0, 'final_totals_after': 0.0,
                'by_house_rooms': {}, 'by_floor': {}, 'by_rooms_comparison': {}
            }

        stats_with = init_stats(total_rate_with_action, percent_change, adjusted_percent)
        stats_no = init_stats(total_rate_no_action, percent_change, adjusted_percent)

        excel_results = []
        house_map = {h.id: h.name for h in mysql_session.query(EstateHouse).filter_by(complex_name=complex_name).all()}

        for obj in all_objects:
            if not obj.estate_price or not obj.estate_area or obj.estate_area <= 0:
                continue

            unit_increase_pct = 0 if obj.id in excluded_ids else adjusted_percent

            # 1. Расчет для технического реестра (Sheet1) — всегда по текущей схеме (с акцией)
            c_net_with = (obj.estate_price - current_res_fee) * mult_with
            c_sqm_usd_with = (c_net_with / obj.estate_area) / usd_rate
            n_sqm_usd_with = c_sqm_usd_with * (1 + unit_increase_pct)
            n_estate_price = ((n_sqm_usd_with * usd_rate * obj.estate_area) / mult_with) + current_res_fee
            excel_results.append({
                'Id обьекта': obj.id,
                'Тип недвижимости': property_type_ru,
                'Новая стоимость': round(n_estate_price, -3)
            })

            # 2. Сбор статистики только для остатков
            if obj.estate_sell_status_name in REMAINDER_STATUSES:
                _fill_stats(obj, stats_with, mult_with, unit_increase_pct, current_res_fee, usd_rate, house_map)
                _fill_stats(obj, stats_no, mult_no, unit_increase_pct, current_res_fee, usd_rate, house_map)

        return excel_results, stats_with, stats_no
    finally:
        mysql_session.close()
        planning_session.close()


def _fill_stats(obj, stats, multiplier, unit_pct, fee, rate, house_map):
    """Вспомогательная функция для наполнения словаря статистики."""
    c_net = (obj.estate_price - fee) * multiplier
    c_sqm_usd = (c_net / obj.estate_area) / rate
    n_sqm_usd = c_sqm_usd * (1 + unit_pct)

    stats['remainder_units'] += 1
    stats['remainder_sqm'] += obj.estate_area
    stats['floor_prices_before'].append(c_sqm_usd)
    stats['floor_prices_after'].append(n_sqm_usd)

    # Стоимость конкретного объекта (чистая)
    c_total_bottom_usd = c_net / rate
    n_total_bottom_usd = (n_sqm_usd * obj.estate_area)

    stats['floor_totals_before'].append(c_total_bottom_usd)
    stats['floor_totals_after'].append(n_total_bottom_usd)

    # ИСПРАВЛЕНО: Теперь суммируем чистую стоимость (Bottom Price), а не грязную
    stats['final_totals_before'] += c_total_bottom_usd
    stats['final_totals_after'] += n_total_bottom_usd

    h_name = house_map.get(obj.house_id, "Неизвестно")
    hr_key = (h_name, obj.estate_rooms or 0)
    stats['by_house_rooms'].setdefault(hr_key, []).append(n_sqm_usd)
    stats['by_floor'].setdefault(obj.estate_floor or 0, []).append(n_sqm_usd)

    rm_cmp = stats['by_rooms_comparison'].setdefault(obj.estate_rooms or 0, {'before': [], 'after': []})
    rm_cmp['before'].append(c_sqm_usd)
    rm_cmp['after'].append(n_sqm_usd)


def _set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def generate_pricelist_excel(results, stats_with, stats_no):
    output = io.BytesIO()
    wb = Workbook()

    # Лист 1: Технический реестр
    ws1 = wb.active
    ws1.title = "PriceList"
    header = ["Id обьекта", "Тип недвижимости", "Новая стоимость"]
    ws1.append(header)
    for res in results:
        ws1.append([res['Id обьекта'], res['Тип недвижимости'], res['Новая стоимость']])
    ws1.auto_filter.ref = ws1.dimensions
    # Лист 2: На подпись (с учетом акций)
    _draw_analytical_sheet(wb.create_sheet("На_подпись"), stats_with)

    # Лист 3: На подпись (без учета акций)
    _draw_analytical_sheet(wb.create_sheet("На_подпись_без_акции"), stats_no)

    wb.save(output)
    output.seek(0)
    return output


def _draw_analytical_sheet(ws, stats):
    """Отрисовка полной структуры аналитической таблицы на листе Excel."""
    f_header = Font(bold=True, name='Arial', size=11)
    f_bold = Font(bold=True, name='Arial', size=10)
    a_center = Alignment(horizontal='center', vertical='center')
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    for col in ['F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col].width = 15

    # 1. Метаданные
    metadata = [
        ("Проект", stats['complex_name']),
        ("Тип недвижимости", stats['prop_type']),
        ("Всего товарного запаса, шт", stats['remainder_units']),
        ("Всего товарного запаса, кв.м", round(stats['remainder_sqm'], 2)),
        ("Сумма регламентных скидок, %", f"{stats['discount_pct']}%"),
        ("Целевой процент повышения, %", f"{stats['percent_change']}%"),
        ("Эффективный % (после размытия), %", f"{stats['adjusted_percent']}%")
    ]

    for i, (label, value) in enumerate(metadata, 2):
        ws.cell(row=i, column=2, value=label).font = f_bold
        ws.cell(row=i, column=2).fill = fill_gray
        ws.cell(row=i, column=3, value=value)
        _set_border(ws, f'B{i}:C{i}')

    # 2. Таблица Изменение цены
    ws.merge_cells('B11:D11')
    hdr = ws['B11']
    hdr.value = "Изменение цены (чистая стоимость дна)"
    hdr.font = f_header
    hdr.alignment = a_center
    hdr.fill = fill_yellow
    _set_border(ws, 'B11:D11')

    ws['C13'], ws['D13'] = "Было", "Стало"
    for col in [3, 4]:
        cell = ws.cell(row=13, column=col)
        cell.font = f_bold
        cell.alignment = a_center
        cell.fill = fill_gray
    _set_border(ws, 'C13:D13')

    # Цена дна
    for idx, label in enumerate(["Минимальная цена дна, $", "Средняя цена дна, $", "Максимальная цена дна, $"]):
        row_idx = 14 + idx
        ws.cell(row=row_idx, column=2, value=label).font = f_bold
        before, after = stats['floor_prices_before'], stats['floor_prices_after']
        if before:
            vals = [min(before), sum(before) / len(before), max(before)]
            n_vals = [min(after), sum(after) / len(after), max(after)]
            ws.cell(row=row_idx, column=3, value=round(vals[idx], 2))
            ws.cell(row=row_idx, column=4, value=round(n_vals[idx], 2))
        _set_border(ws, f'B{row_idx}:D{row_idx}')

    # Стоимость дна
    for idx, label in enumerate(
            ["Минимальная стоимость дна, $", "Средняя стоимость дна, $", "Максимальная стоимость дна, $"]):
        row_idx = 18 + idx
        ws.cell(row=row_idx, column=2, value=label).font = f_bold
        before, after = stats['floor_totals_before'], stats['floor_totals_after']
        if before:
            vals = [min(before), sum(before) / len(before), max(before)]
            n_vals = [min(after), sum(after) / len(after), max(after)]
            ws.cell(row=row_idx, column=3, value=round(vals[idx], 0))
            ws.cell(row=row_idx, column=4, value=round(n_vals[idx], 0))
        _set_border(ws, f'B{row_idx}:D{row_idx}')

    # Общая стоимость (Bottom Price)
    ws.cell(row=22, column=2, value="Общая чистая стоимость остатков, $").font = f_header
    ws.cell(row=22, column=3, value=round(stats['final_totals_before'], 0))
    ws.cell(row=22, column=4, value=round(stats['final_totals_after'], 0))
    _set_border(ws, 'B22:D22')

    # 3. Таблица типологии (Было/Стало)
    start_row_typ = 25
    ws.merge_cells(f'B{start_row_typ}:H{start_row_typ}')
    typ_hdr = ws.cell(row=start_row_typ, column=2, value="Сравнение по типологии (цена дна за кв.м, $)")
    typ_hdr.font = f_header
    typ_hdr.alignment = a_center
    typ_hdr.fill = fill_yellow
    _set_border(ws, f'B{start_row_typ}:H{start_row_typ}')

    h_row = start_row_typ + 1
    sub_headers = ["Комн.", "Мин (Было)", "Мин (Стало)", "Сред (Было)", "Сред (Стало)", "Макс (Было)", "Макс (Стало)"]
    for i, text in enumerate(sub_headers):
        cell = ws.cell(row=h_row, column=2 + i, value=text)
        cell.font = f_bold
        cell.fill = fill_gray
        cell.alignment = a_center
    _set_border(ws, f'B{h_row}:H{h_row}')

    curr_row = h_row + 1
    for rooms in sorted(stats['by_rooms_comparison'].keys()):
        data = stats['by_rooms_comparison'][rooms]
        b_vals, a_vals = data['before'], data['after']
        ws.cell(row=curr_row, column=2, value=f"{rooms}-комн")
        ws.cell(row=curr_row, column=3, value=round(min(b_vals), 2))
        ws.cell(row=curr_row, column=4, value=round(min(a_vals), 2))
        ws.cell(row=curr_row, column=5, value=round(sum(b_vals) / len(b_vals), 2))
        ws.cell(row=curr_row, column=6, value=round(sum(a_vals) / len(a_vals), 2))
        ws.cell(row=curr_row, column=7, value=round(max(b_vals), 2))
        ws.cell(row=curr_row, column=8, value=round(max(a_vals), 2))
        _set_border(ws, f'B{curr_row}:H{curr_row}')
        curr_row += 1

    # 4. Таблицы справа
    row = 2
    for i, h in enumerate(["Дом", "Комн.", "Мин. цена $", "Сред. цена $", "Макс. цена $"]):
        cell = ws.cell(row=row, column=6 + i, value=h)
        cell.font, cell.fill, cell.alignment = f_bold, fill_gray, a_center
    _set_border(ws, f'F{row}:J{row}')

    row += 1
    for (h_name, rooms), prices in sorted(stats['by_house_rooms'].items()):
        ws.cell(row=row, column=6, value=h_name)
        ws.cell(row=row, column=7, value=rooms)
        ws.cell(row=row, column=8, value=round(min(prices), 2))
        ws.cell(row=row, column=9, value=round(sum(prices) / len(prices), 2))
        ws.cell(row=row, column=10, value=round(max(prices), 2))
        _set_border(ws, f'F{row}:J{row}')
        row += 1

    row_fl = 2
    for i, h in enumerate(["Этаж", "Мин. цена $", "Сред. цена $", "Макс. цена $"]):
        cell = ws.cell(row=row_fl, column=12 + i, value=h)
        cell.font, cell.fill, cell.alignment = f_bold, fill_gray, a_center
    _set_border(ws, f'L{row_fl}:O{row_fl}')

    row_fl += 1
    for floor, prices in sorted(stats['by_floor'].items()):
        ws.cell(row=row_fl, column=12, value=floor)
        ws.cell(row=row_fl, column=13, value=round(min(prices), 2))
        ws.cell(row=row_fl, column=14, value=round(sum(prices) / len(prices), 2))
        ws.cell(row=row_fl, column=15, value=round(max(prices), 2))
        _set_border(ws, f'L{row_fl}:O{row_fl}')
        row_fl += 1